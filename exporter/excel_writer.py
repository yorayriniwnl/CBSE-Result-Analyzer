"""
Excel Writer: produces a 4-sheet Excel workbook.
  Sheet 1 — Student Data
  Sheet 2 — Subject Analysis
  Sheet 3 — School Summary
  Sheet 4 — Parse Errors (if any)
"""

from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from parser.gazette_parser import ParseError, Student
from transformer.calculator import compute_subject_analysis, compute_summary
from transformer.normalizer import build_normalized_table


# ── Colour palette ───────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # dark navy
CLR_HEADER_FG   = "FFFFFF"   # white
CLR_PASS        = "1A5C38"   # dark green
CLR_FAIL        = "C00000"   # deep red
CLR_COMP        = "7F4C00"   # amber
CLR_ABSENT      = "595959"   # grey
CLR_ALT_ROW     = "EBF2FF"   # very light blue
CLR_EMPTY_CELL  = "F2F2F2"   # light grey
CLR_GOOD_AVG    = "C6EFCE"   # light green
CLR_BAD_AVG     = "FFC7CE"   # light red
CLR_MID_AVG     = "FFEB9C"   # yellow

THIN = Side(style='thin', color="B0B0B0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cv = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(cv) + 2)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ── Sheet 1: Student Data ────────────────────────────────────────────────────
def _write_student_sheet(wb: Workbook, df: pd.DataFrame, title: str):
    ws = wb.active
    ws.title = "Student Data"

    # School title row
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(df.columns))
    ws['A1'].font = Font(name="Calibri", bold=True, size=13, color=CLR_HEADER_FG)
    ws['A1'].fill = _hdr_fill(CLR_HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Header row
    headers = list(df.columns)
    ws.append(headers)
    hdr_row = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill = _hdr_fill(CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 32

    # Result column index
    result_col = headers.index('Result') + 1 if 'Result' in headers else None
    # Subject mark columns (anything that ends with a code pattern like '(NNN)')
    import re
    mark_cols = [i + 1 for i, h in enumerate(headers)
                 if re.search(r'\(\d{3}\)$', str(h))]

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        alt = (r_idx % 2 == 0)
        for c_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell = ws.cell(row=r_idx, column=c_idx)

            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = None
                if c_idx in mark_cols:
                    cell.fill = _hdr_fill(CLR_EMPTY_CELL)
            else:
                cell.value = val

            cell.border = THIN_BORDER
            cell.font = _font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Alternate row shading
            if alt and cell.fill.fgColor.rgb in ('00000000', 'FFFFFFFF', '00FFFFFF'):
                cell.fill = _hdr_fill(CLR_ALT_ROW)

        # Result colour
        if result_col:
            rc = ws.cell(row=r_idx, column=result_col)
            result_val = str(row.get('Result', '')).upper()
            if result_val == 'PASS':
                rc.font = _font(bold=True, color=CLR_PASS, size=10)
            elif result_val == 'FAIL':
                rc.font = _font(bold=True, color=CLR_FAIL, size=10)
            elif result_val == 'COMP':
                rc.font = _font(bold=True, color=CLR_COMP, size=10)

    # Freeze top 2 rows + first column
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = ws.dimensions

    _auto_width(ws)
    # Name column wider
    if 'Name' in headers:
        ws.column_dimensions[get_column_letter(headers.index('Name') + 1)].width = 28


# ── Sheet 2: Subject Analysis ────────────────────────────────────────────────
def _write_analysis_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Subject Analysis")

    headers = list(df.columns)
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _font(bold=True, color=CLR_HEADER_FG)
        cell.fill = _hdr_fill(CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    avg_col = headers.index('Average Marks') + 1 if 'Average Marks' in headers else None

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val if not (isinstance(val, float) and pd.isna(val)) else None
            cell.border = THIN_BORDER
            cell.font = _font(size=10)
            cell.alignment = Alignment(horizontal='center')

        # Colour-code average
        if avg_col:
            avg = row.get('Average Marks', None)
            ac = ws.cell(row=r_idx, column=avg_col)
            if avg is not None and not (isinstance(avg, float) and pd.isna(avg)):
                if avg >= 75:
                    ac.fill = _hdr_fill(CLR_GOOD_AVG)
                elif avg < 40:
                    ac.fill = _hdr_fill(CLR_BAD_AVG)
                else:
                    ac.fill = _hdr_fill(CLR_MID_AVG)

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 3: School Summary ──────────────────────────────────────────────────
def _write_summary_sheet(wb: Workbook, summary: Dict, title: str):
    ws = wb.create_sheet("School Summary")

    ws.append([title])
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(name="Calibri", bold=True, size=13, color=CLR_HEADER_FG)
    ws['A1'].fill = _hdr_fill(CLR_HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    ws.append(['Metric', 'Value'])
    for c in range(1, 3):
        cell = ws.cell(row=2, column=c)
        cell.font = _font(bold=True, color=CLR_HEADER_FG)
        cell.fill = _hdr_fill(CLR_HEADER_BG)
        cell.border = THIN_BORDER

    for r, (k, v) in enumerate(summary.items(), 3):
        ws.cell(row=r, column=1, value=k).border = THIN_BORDER
        ws.cell(row=r, column=1).font = _font(bold=True)
        vc = ws.cell(row=r, column=2, value=v)
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal='center')

    _auto_width(ws, min_width=18)


# ── Sheet 4: Parse Errors ────────────────────────────────────────────────────
def _write_errors_sheet(wb: Workbook, errors: List[ParseError]):
    ws = wb.create_sheet("Parse Errors")
    headers = ['Level', 'Roll No', 'Line No', 'Message']
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _font(bold=True, color=CLR_HEADER_FG)
        cell.fill = _hdr_fill("7F0000" if errors else "1F4E79")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    if not errors:
        ws.append(['✅ No parse errors detected', '', '', ''])
    else:
        for r, err in enumerate(errors, 2):
            row_data = [err.level, err.roll, err.line_no, err.message]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = THIN_BORDER
                if err.level == 'ERROR':
                    cell.fill = _hdr_fill(CLR_BAD_AVG)
                elif err.level == 'WARNING':
                    cell.fill = _hdr_fill(CLR_MID_AVG)

    _auto_width(ws)


# ── Public entry point ───────────────────────────────────────────────────────
def write_excel(
    students: List[Student],
    errors: List[ParseError],
    subject_master: Dict,
    output_path: str,
    school_name: str = "CBSE Results 2026"
):
    df_students, all_codes = build_normalized_table(students, subject_master)
    df_analysis = compute_subject_analysis(students, all_codes, subject_master)
    summary = compute_summary(students)

    wb = Workbook()

    _write_student_sheet(wb, df_students, school_name)
    _write_analysis_sheet(wb, df_analysis)
    _write_summary_sheet(wb, summary, school_name)
    _write_errors_sheet(wb, errors)

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")
    print(f"   Students : {len(students)}")
    print(f"   Subjects : {len(all_codes)}")
    print(f"   Errors   : {len(errors)}")
